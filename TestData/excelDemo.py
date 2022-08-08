import openpyxl
book = openpyxl.load_workbook("PythonDemo.xlsx")
sheet = book.active
Dict = {}
cell = sheet.cell(row=3, column=5)
print(cell.value)
sheet.cell(row=3, column=5).value = "Artem"
print(sheet.cell(row=3, column=5).value)
print(sheet.max_row)
print(sheet.max_column)
print(sheet["D134"].value)

for i in range(1, sheet.max_row):
    for j in range(1, sheet.max_column+1):
        if sheet.cell(row=i, column=j).value == "Top 20":
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            # print(sheet.cell(row=i, column=j).value)

print(Dict)